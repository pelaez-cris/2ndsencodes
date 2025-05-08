from openpyxl import load_workbook

workbook = load_workbook('sample.xlsx')
sheet = workbook.active

print('Reading Excel Files;/n')
for row in sheet.iter_rows(values_only=True):
    print(row)